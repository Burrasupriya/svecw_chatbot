# import os
# from dotenv import load_dotenv
# import google.generativeai as genai
# from langchain_huggingface import HuggingFaceEmbeddings
# from langchain_community.vectorstores import FAISS
# from langchain.text_splitter import CharacterTextSplitter
# from langchain.docstore.document import Document
# from PyPDF2 import PdfReader


# load_dotenv()

# # Configure Google Gemini API
# GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
# genai.configure(api_key=GOOGLE_API_KEY)
# model = genai.GenerativeModel("gemini-1.5-flash")

# # Initialize Hugging Face Embeddings with explicit model name
# embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")

# # Load and Split Document using PyPDF2
# def load_document(file_path):
#     documents = []
#     pdf = PdfReader(file_path)
#     text_content = ""
#     for page in pdf.pages:
#         text_content += page.extract_text()
    
#     text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
#     chunks = text_splitter.split_text(text_content)
#     for chunk in chunks:
#         documents.append(Document(page_content=chunk))
#     return documents

# # Create Vector Store
# def create_vector_store(documents):
#     vector_store = FAISS.from_documents(documents, embeddings)
#     return vector_store

# # Load the single document and create vector store
# file_path = './data/svecw2024.pdf'  # Path to your PDF file
# documents = load_document(file_path)
# vector_store = create_vector_store(documents)

# # Function to get chatbot response using Gemini and LangChain
# def get_response(prompt):
#     # Retrieve the most relevant document chunk
#     search_results = vector_store.similarity_search(prompt, k=1)
#     if search_results:
#         context = search_results[0].page_content
#         # Include context for a more accurate response
#         full_prompt = f"Context: {context}\nUser: {prompt}\n\Bot:"
#     else:
#         full_prompt = f"User: {prompt}\nBot:"
    
#     # Generate response using Gemini model
#     response = model.generate_content(full_prompt)
#     return response.candidates[0].content.parts[0].text

# if __name__ == "__main__":
#     while True:
#         user_input = input("You: ")
#         if user_input.lower() == "exit":
#             print("Exiting the chatbot. Goodbye!")
#             break
        
#         bot_response = get_response(user_input)
#         print(f"Bot: {bot_response}")



import os
from dotenv import load_dotenv
import google.generativeai as genai
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from langchain.docstore.document import Document
from openpyxl import load_workbook  # For reading XLSX files

load_dotenv()

# Configure Google Gemini API
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=GOOGLE_API_KEY)
model = genai.GenerativeModel("gemini-1.5-flash")

# Initialize Hugging Face Embeddings with explicit model name
embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")

# Load and Split Document using openpyxl for XLSX files
def load_document(file_path):
    documents = []
    workbook = load_workbook(file_path)
    text_content = ""
    
    # Iterate through each sheet and each row
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text_content += str(cell) + " "  # Concatenate each cell's text
    
    text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
    chunks = text_splitter.split_text(text_content)
    for chunk in chunks:
        documents.append(Document(page_content=chunk))
    
    return documents

# Create Vector Store
def create_vector_store(documents):
    vector_store = FAISS.from_documents(documents, embeddings)
    return vector_store

# Load the single document and create vector store
file_path = './data/svecw2024.xlsx'  # Path to your XLSX file
documents = load_document(file_path)
vector_store = create_vector_store(documents)

# Function to get chatbot response using Gemini and LangChain
def get_response(prompt):
    # Retrieve the most relevant document chunk
    search_results = vector_store.similarity_search(prompt, k=1)
    if search_results:
        context = search_results[0].page_content
        # Include context for a more accurate response
        full_prompt = f"Context: {context}\nUser: {prompt}\nBot:"
    else:
        full_prompt = f"User: {prompt}\nBot:"
    
    # Generate response using Gemini model
    response = model.generate_content(full_prompt)
    return response.candidates[0].content.parts[0].text

if __name__ == "__main__":
    while True:
        user_input = input("You: ")
        if user_input.lower() == "exit":
            print("Exiting the chatbot. Goodbye!")
            break
        
        bot_response = get_response(user_input)
        print(f"Bot: {bot_response}")
